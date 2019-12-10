from openpyxl import load_workbook
wb=load_workbook("测试哦.xlsx")
sheet = wb["1"]
date1=[]
for j in range(1,sheet.max_row+1):
    for i in range(1,sheet.max_column+1):
        a=(j,i)
        b=sheet.cell(j,i).value
        date = {}
        date[a]=b
        date1.append(date)
    print(date1)
    date1 = []

